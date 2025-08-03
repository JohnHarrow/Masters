import pandas as pd
from collections import deque
from pulp import LpProblem, LpVariable, LpMaximize, lpSum, LpBinary, PULP_CBC_CMD
import numpy as np
import matplotlib.pyplot as plt
from openpyxl.styles import PatternFill, Font

# --- Load data ---
students_df = pd.read_csv("students.csv")
projects_df = pd.read_csv("projects.csv")
preallocated_df = pd.read_csv("preallocated.csv")
supervisors_df = pd.read_csv("supervisors.csv")  # supervisor_id, supervisor_name, capacity

# --- Validate student data structure and content ---
def validate_student_data(students_df, projects_df):
    errors = []
    required_fields = ['student_id', 'student_name', 'choice_1', 'choice_2', 'choice_3']
    missing_fields = [field for field in required_fields if field not in students_df.columns]
    if missing_fields:
        errors.append(f"Missing required columns: {missing_fields}")

    if students_df[required_fields].isnull().any().any():
        null_rows = students_df[required_fields].isnull().any(axis=1)
        null_ids = students_df[null_rows]['student_id'].tolist()
        errors.append(f"Missing data in required fields for student_ids: {null_ids}")

    if students_df['student_id'].duplicated().any():
        dup_ids = students_df[students_df['student_id'].duplicated()]['student_id'].tolist()
        errors.append(f"Duplicate student IDs found: {dup_ids}")

    valid_project_ids = set(projects_df['project_id'])
    for _, row in students_df.iterrows():
        sid = row['student_id']
        for col in ['choice_1', 'choice_2', 'choice_3']:
            if row[col] not in valid_project_ids:
                errors.append(f"Student {sid}: Invalid project ID '{row[col]}' in {col}")

    for _, row in students_df.iterrows():
        sid = row['student_id']
        choices = [row['choice_1'], row['choice_2'], row['choice_3']]
        if len(set(choices)) < 3:
            errors.append(f"Student {sid}: Duplicate project choices detected.")

    if errors:
        print("❌ STUDENT DATA VALIDATION ERRORS:")
        for err in errors:
            print(" -", err)
        raise ValueError("Student data validation failed.")
    else:
        print("✅ Student data validation passed.")

# --- Optional supervisor diversity check ---
def validate_supervisor_diversity(students_df, projects_df):
    warnings = []
    for _, row in students_df.iterrows():
        choices = [row['choice_1'], row['choice_2'], row['choice_3']]
        sup_ids = []
        for c in choices:
            pr = projects_df[projects_df['project_id'] == c]
            if not pr.empty:
                sup_ids.append(pr['supervisor_id'].values[0])
        if len(set(sup_ids)) < 3:
            warnings.append(row['student_id'])

    if warnings:
        print("⚠️ SUPERVISOR DIVERSITY WARNINGS (not errors):")
        print(" Students with multiple choices from same supervisor:", warnings)
    else:
        print("✅ Supervisor diversity check passed.")

# --- Capacities ---
supervisor_capacity = {
    row['supervisor_id']: int(row['capacity']) if pd.notna(row['capacity']) else 3
    for _, row in supervisors_df.iterrows()
}

project_capacity = {
    row['project_id']: int(row['max_students']) if pd.notna(row['max_students']) else None
    for _, row in projects_df.iterrows()
}

# --- Validations ---
validate_student_data(students_df, projects_df)
validate_supervisor_diversity(students_df, projects_df)

# --- Greedy matching function ---
def greedy_matching(students_df, projects_df, supervisor_capacity, project_capacity, preallocated_df):
    allocation = {}
    supervisor_load = {sup_id: 0 for sup_id in supervisor_capacity}
    project_load = {pid: 0 for pid in projects_df['project_id']}

    # Assign preallocated students
    for _, row in preallocated_df.iterrows():
        sid = row['student_id']
        pid = row['project_id']
        sup = row['supervisor_id']
        allocation[sid] = pid
        supervisor_load[sup] += 1
        project_load[pid] += 1

    # 🔀 Shuffle the students to reduce ordering bias
    students_df_shuffled = students_df.sample(frac=1, random_state=None).reset_index(drop=True)

    # Greedy allocation based on student preferences
    for _, row in students_df_shuffled.iterrows():
        sid = row['student_id']
        if sid in allocation:
            continue

        for choice in ['choice_1', 'choice_2', 'choice_3']:
            pid = row[choice]
            project_row = projects_df[projects_df['project_id'] == pid]
            if project_row.empty:
                continue

            sup = project_row['supervisor_id'].values[0]
            max_proj_cap = project_capacity[pid]

            if supervisor_load[sup] < supervisor_capacity[sup]:
                if max_proj_cap is None or project_load[pid] < max_proj_cap:
                    allocation[sid] = pid
                    supervisor_load[sup] += 1
                    project_load[pid] += 1
                    break

    return allocation

# --- Stable marriage matching function ---
def stable_marriage_matching(students_df, projects_df, supervisor_capacity, project_capacity, preallocated_df):
    allocation = {}
    supervisor_load = {sup_id: 0 for sup_id in supervisor_capacity}
    project_load = {pid: 0 for pid in projects_df['project_id']}

    for _, row in preallocated_df.iterrows():
        sid = row['student_id']
        pid = row['project_id']
        sup = row['supervisor_id']
        allocation[sid] = pid
        supervisor_load[sup] += 1
        project_load[pid] += 1

    student_prefs = {
        row['student_id']: deque([row['choice_1'], row['choice_2'], row['choice_3']])
        for _, row in students_df.iterrows() if row['student_id'] not in allocation
    }

    free_students = deque(student_prefs.keys())

    while free_students:
        sid = free_students.popleft()
        if not student_prefs[sid]:
            continue

        pid = student_prefs[sid].popleft()
        project_row = projects_df[projects_df['project_id'] == pid]
        if project_row.empty:
            continue

        sup = project_row['supervisor_id'].values[0]
        max_proj_cap = project_capacity.get(pid, None)

        if supervisor_load[sup] < supervisor_capacity[sup] and \
           (max_proj_cap is None or project_load[pid] < max_proj_cap):
            allocation[sid] = pid
            supervisor_load[sup] += 1
            project_load[pid] += 1
        else:
            free_students.append(sid)

    return allocation

# --- Linear programming matching function ---
def linear_programming_matching(students_df, projects_df, supervisor_capacity, project_capacity, preallocated_df):
    prob = LpProblem("Student_Project_Matching", LpMaximize)

    students = list(students_df['student_id'])
    projects = list(projects_df['project_id'])

    # Build dictionaries for quick lookup
    project_supervisors = projects_df.set_index('project_id')['supervisor_id'].to_dict()
    student_choices = {}
    for _, row in students_df.iterrows():
        sid = row['student_id']
        # Map project to weight based on preference order (choice_1=3, choice_2=2, choice_3=1)
        student_choices[sid] = {
            row['choice_1']: 3,
            row['choice_2']: 2,
            row['choice_3']: 1
        }

    # Decision variables: x[(student, project)] = 1 if assigned, else 0
    x = LpVariable.dicts("assign",
                         [(s, p) for s in students for p in projects],
                         cat=LpBinary)

    # Objective: maximize total satisfaction (sum of assigned weights)
    prob += lpSum(x[(s, p)] * student_choices.get(s, {}).get(p, 0) for s in students for p in projects)

    # Constraint 1: Each student assigned to at most one project
    for s in students:
        prob += lpSum(x[(s, p)] for p in projects) <= 1

    # Constraint 2: Respect preallocation (force assigned)
    for _, row in preallocated_df.iterrows():
        sid = row['student_id']
        pid = row['project_id']
        for p in projects:
            if p == pid:
                prob += x[(sid, p)] == 1
            else:
                prob += x[(sid, p)] == 0

    # Constraint 3: Project capacity
    for p in projects:
        max_cap = project_capacity.get(p)
        if max_cap is not None:
            prob += lpSum(x[(s, p)] for s in students) <= max_cap

    # Constraint 4: Supervisor capacity
    for sup, sup_cap in supervisor_capacity.items():
        sup_projects = [p for p, sup_id in project_supervisors.items() if sup_id == sup]
        prob += lpSum(x[(s, p)] for s in students for p in sup_projects) <= sup_cap

    # Solve
    solver = PULP_CBC_CMD(msg=False)
    prob.solve(solver)

    # Extract allocation
    allocation = {}
    for s in students:
        for p in projects:
            var = x[(s, p)]
            if var.varValue is not None and var.varValue > 0.5:
                allocation[s] = p
                break

    return allocation

# --- Save results ---
def save_allocation_result(allocation, students_df, filename):
    output = []
    for _, row in students_df.iterrows():
        sid = row['student_id']
        assigned = allocation.get(sid, "UNASSIGNED")
        output.append({
            "student_id": sid,
            "student_name": row['student_name'],
            "assigned_project": assigned
        })
    pd.DataFrame(output).to_csv(filename, index=False)
    print(f"📄 Results saved to {filename}")

# --- New: Analyze match quality ---
def analyze_match_quality(allocation, students_df, projects_df, supervisors_df, method_name="Method"):
    print(f"\n📊 Analysis of match quality for {method_name}:")

    # 1. Choice preference distribution
    distribution = {'choice_1': 0, 'choice_2': 0, 'choice_3': 0, 'unmatched': 0, 'other': 0}
    for _, row in students_df.iterrows():
        sid = row['student_id']
        assigned = allocation.get(sid, None)
        if assigned is None:
            distribution['unmatched'] += 1
        elif assigned == row['choice_1']:
            distribution['choice_1'] += 1
        elif assigned == row['choice_2']:
            distribution['choice_2'] += 1
        elif assigned == row['choice_3']:
            distribution['choice_3'] += 1
        else:
            distribution['other'] += 1  # likely preallocated or forced assignments

    print("Choice preference distribution:")
    for choice, count in distribution.items():
        print(f"  {choice}: {count}")

    # 2. Supervisor load distribution
    proj_to_sup = projects_df.set_index('project_id')['supervisor_id'].to_dict()
    supervisor_load = {sup: 0 for sup in supervisors_df['supervisor_id']}
    for sid, pid in allocation.items():
        sup = proj_to_sup.get(pid)
        if sup is not None:
            supervisor_load[sup] += 1

    loads = np.array(list(supervisor_load.values()))
    print("\nSupervisor load distribution:")
    print(f"  Min load: {loads.min()}")
    print(f"  Max load: {loads.max()}")
    print(f"  Mean load: {loads.mean():.2f}")
    print(f"  Std dev load: {loads.std():.2f}")

    # 3. Students assigned outside their choices
    outside = []
    for _, row in students_df.iterrows():
        sid = row['student_id']
        assigned = allocation.get(sid, None)
        if assigned is not None and assigned not in [row['choice_1'], row['choice_2'], row['choice_3']]:
            outside.append(sid)
    print(f"\nStudents assigned outside their choices: {len(outside)}")
    if outside:
        print("  Student IDs:", outside)

    # 4. Project utilization
    usage = {pid: 0 for pid in projects_df['project_id']}
    for pid in allocation.values():
        usage[pid] = usage.get(pid, 0) + 1

    print("\nProject utilization (project_id: number assigned):")
    for pid, count in usage.items():
        print(f"  {pid}: {count}")

# --- Satisfaction analysis ---
def compute_satisfaction_scores(allocation, students_df, method_name="Method"):
    print(f"\n🎯 Student Satisfaction Score for {method_name}")

    score_weights = {'choice_1': 3, 'choice_2': 2, 'choice_3': 1}
    scores = []
    unmatched = 0

    for _, row in students_df.iterrows():
        sid = row['student_id']
        assigned = allocation.get(sid)

        if assigned == row['choice_1']:
            scores.append(score_weights['choice_1'])
        elif assigned == row['choice_2']:
            scores.append(score_weights['choice_2'])
        elif assigned == row['choice_3']:
            scores.append(score_weights['choice_3'])
        elif assigned is None:
            unmatched += 1
            scores.append(0)
        else:
            scores.append(0)

    avg_score = np.mean(scores)
    print(f"  Average satisfaction score: {avg_score:.2f}")
    print(f"  Unmatched students: {unmatched}")

    # Histogram plot
    plt.figure(figsize=(6, 4))
    plt.hist(scores, bins=[-0.5, 0.5, 1.5, 2.5, 3.5], edgecolor='black', align='mid', rwidth=0.8)
    plt.xticks([0, 1, 2, 3])
    plt.xlabel("Satisfaction Score (0–3)")
    plt.ylabel("Number of Students")
    plt.title(f"Satisfaction Score Distribution - {method_name}")
    plt.grid(axis='y', linestyle='--', alpha=0.7)
    plt.tight_layout()
    plt.show()

# --- Supervisor load analysis ---
def analyze_supervisor_load(allocation, projects_df, supervisors_df, method_name="Method"):
    print(f"\n📋 Supervisor Load Analysis for {method_name}:")

    proj_to_sup = projects_df.set_index('project_id')['supervisor_id'].to_dict()
    sup_capacities = supervisors_df.set_index('supervisor_id')['capacity'].to_dict()

    supervisor_load = {sup: 0 for sup in sup_capacities}

    # Count how many students are assigned under each supervisor
    for pid in allocation.values():
        sup = proj_to_sup.get(pid)
        if sup in supervisor_load:
            supervisor_load[sup] += 1

    # Print and highlight overloads or underutilized supervisors
    for sup_id, load in supervisor_load.items():
        cap = sup_capacities.get(sup_id, 0)
        sup_name = supervisors_df[supervisors_df['supervisor_id'] == sup_id]['supervisor_name'].values[0]

        status = ""
        if load > cap:
            status = "❌ OVERLOADED"
        elif load < cap:
            status = "⚠️ UNDERUSED"
        else:
            status = "✅ OPTIMAL"

        print(f"  {sup_name} (ID: {sup_id}) – Load: {load}, Capacity: {cap} → {status}")


# --- Project popularity analysis ---
def analyze_project_popularity_and_utilization(students_df, projects_df, allocation, method_name="Method"):
    print(f"\n📈 Project Popularity & Utilization – {method_name}")

    # 1. Count how many students requested each project (1st, 2nd, or 3rd choice)
    popularity = {pid: 0 for pid in projects_df['project_id']}
    for _, row in students_df.iterrows():
        for col in ['choice_1', 'choice_2', 'choice_3']:
            pid = row[col]
            if pid in popularity:
                popularity[pid] += 1

    # 2. Count how many students were actually assigned to each project
    utilization = {pid: 0 for pid in projects_df['project_id']}
    for pid in allocation.values():
        if pid in utilization:
            utilization[pid] += 1

    # 3. Report
    print("\nProject ID | Requested | Assigned | Status")
    print("-" * 40)
    for pid in projects_df['project_id']:
        requested = popularity.get(pid, 0)
        assigned = utilization.get(pid, 0)

        if requested == 0:
            status = "🧊 NEVER PICKED"
        elif assigned == 0:
            status = "⚠️ NEVER ASSIGNED"
        elif assigned > requested:
            status = "❌ OVER-ASSIGNED"
        elif assigned < requested:
            status = "🔼 UNDER-ASSIGNED"
        else:
            status = "✅ MATCHED"

        print(f"{pid:<10} | {requested:<9} | {assigned:<8} | {status}")

def export_matching_results_to_excel(
    students_df,
    allocations_dict,
    filename="matching_results.xlsx"
):
    """
    Exports matching results for multiple algorithms into one Excel workbook.
    
    Parameters:
    - students_df: DataFrame with columns ['student_id', 'choice_1', 'choice_2', 'choice_3']
    - allocations_dict: dict of {method_name: allocation_dict} where allocation_dict maps student_id to project_id
    - filename: output Excel filename
    
    Output:
    - Excel workbook with sheets for each algorithm and a summary sheet.
    """
    
    from openpyxl.styles import PatternFill
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl import load_workbook
    
    score_weights = {'choice_1': 3, 'choice_2': 2, 'choice_3': 1}
    choice_columns = ['choice_1', 'choice_2', 'choice_3']

    # Use context manager to save Excel file properly
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:

        # Store summary data for each method
        summary_data = []

        for method_name, allocation in allocations_dict.items():
            # Build DataFrame of results for this method
            results = []
            for _, student in students_df.iterrows():
                sid = student['student_id']
                assigned = allocation.get(sid)

                if assigned is None:
                    choice_num = 'unmatched'
                    score = 0
                else:
                    # Determine which choice number the assigned project corresponds to
                    if assigned == student['choice_1']:
                        choice_num = 1
                        score = score_weights['choice_1']
                    elif assigned == student['choice_2']:
                        choice_num = 2
                        score = score_weights['choice_2']
                    elif assigned == student['choice_3']:
                        choice_num = 3
                        score = score_weights['choice_3']
                    else:
                        choice_num = 'other'
                        score = 0

                results.append({
                    'student_id': sid,
                    'assigned_project': assigned,
                    'assigned_choice': choice_num,
                    'satisfaction_score': score
                })

            results_df = pd.DataFrame(results)

            # Write sheet
            results_df.to_excel(writer, sheet_name=method_name, index=False)

            # Aggregate stats for summary
            total_students = len(students_df)
            unmatched = sum(results_df['assigned_choice'] == 'unmatched')
            avg_score = results_df['satisfaction_score'].mean()
            choice_1_count = sum(results_df['assigned_choice'] == 1)
            choice_2_count = sum(results_df['assigned_choice'] == 2)
            choice_3_count = sum(results_df['assigned_choice'] == 3)
            other_count = sum(results_df['assigned_choice'] == 'other')

            summary_data.append({
                'Algorithm': method_name,
                'Total Students': total_students,
                'Unmatched Students': unmatched,
                'Average Satisfaction Score': avg_score,
                'Assigned Choice 1': choice_1_count,
                'Assigned Choice 2': choice_2_count,
                'Assigned Choice 3': choice_3_count,
                'Assigned Other': other_count
            })

        # Create summary sheet
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

    # Now add conditional formatting using openpyxl (file is fully saved now)
    wb = load_workbook(filename)

    # Define fills
    fill_unmatched = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # light red
    fill_other = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")      # light yellow

    for method_name in allocations_dict.keys():
        ws = wb[method_name]

        # Find the column index of assigned_choice (1-based for openpyxl)
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == 'assigned_choice':
                col_idx = idx
                break

        col_letter = chr(64 + col_idx)
        ws.conditional_formatting.add(
            f'{col_letter}2:{col_letter}{ws.max_row}',
            CellIsRule(operator='equal', formula=['"unmatched"'], fill=fill_unmatched)
        )
        ws.conditional_formatting.add(
            f'{col_letter}2:{col_letter}{ws.max_row}',
            CellIsRule(operator='equal', formula=['"other"'], fill=fill_other)
        )

    wb.save(filename)
    print(f"Matching results successfully saved to {filename}")




# --- Run all matching algorithms ---
allocation_greedy = greedy_matching(
    students_df, projects_df, supervisor_capacity, project_capacity, preallocated_df
)
allocation_stable = stable_marriage_matching(
    students_df, projects_df, supervisor_capacity, project_capacity, preallocated_df
)
allocation_lp = linear_programming_matching(
    students_df, projects_df, supervisor_capacity, project_capacity, preallocated_df
)

# --- Save output ---
save_allocation_result(allocation_greedy, students_df, "allocation_result_greedy.csv")
save_allocation_result(allocation_stable, students_df, "allocation_result_stable.csv")
save_allocation_result(allocation_lp, students_df, "allocation_result_linear_programming.csv")


# --- Summary ---
unmatched_greedy = [sid for sid in students_df['student_id'] if sid not in allocation_greedy]
unmatched_stable = [sid for sid in students_df['student_id'] if sid not in allocation_stable]
unmatched_lp = [sid for sid in students_df['student_id'] if sid not in allocation_lp]

print(f"\n🔁 Greedy unmatched students: {len(unmatched_greedy)}")
print(f"💍 Stable unmatched students: {len(unmatched_stable)}")
print(f"🧮 LP unmatched students: {len(unmatched_lp)}")

# --- Excel Output ---
allocations_dict = {
    "Greedy Matching": allocation_greedy,
    "Stable Marriage": allocation_stable,
    "Linear Programming": allocation_lp,
}

export_matching_results_to_excel(students_df, allocations_dict, "all_matchings.xlsx")

# --- Run analysis on each matching ---
analyze_match_quality(allocation_greedy, students_df, projects_df, supervisors_df, "Greedy Matching")
analyze_match_quality(allocation_stable, students_df, projects_df, supervisors_df, "Stable Marriage Matching")
analyze_match_quality(allocation_lp, students_df, projects_df, supervisors_df, "Linear Programming Matching")

# --- Run satisfaction analysis ---
compute_satisfaction_scores(allocation_greedy, students_df, "Greedy Matching")
compute_satisfaction_scores(allocation_stable, students_df, "Stable Marriage Matching")
compute_satisfaction_scores(allocation_lp, students_df, "Linear Programming Matching")

# --- Run supervisor load analysis ---
analyze_supervisor_load(allocation_greedy, projects_df, supervisors_df, "Greedy Matching")
analyze_supervisor_load(allocation_stable, projects_df, supervisors_df, "Stable Marriage Matching")
analyze_supervisor_load(allocation_lp, projects_df, supervisors_df, "Linear Programming Matching")

# --- Run project popularity analysis ---
analyze_project_popularity_and_utilization(students_df, projects_df, allocation_greedy, "Greedy Matching")
analyze_project_popularity_and_utilization(students_df, projects_df, allocation_stable, "Stable Marriage Matching")
analyze_project_popularity_and_utilization(students_df, projects_df, allocation_lp, "Linear Programming Matching")
