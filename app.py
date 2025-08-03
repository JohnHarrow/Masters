# python -m streamlit run app.py

# ═══════════════════════════════════════════════════════
# SECTION: App Configuration
# ═══════════════════════════════════════════════════════

import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from collections import deque
from pulp import LpProblem, LpVariable, LpMaximize, lpSum, LpBinary, PULP_CBC_CMD
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from tempfile import NamedTemporaryFile
from openpyxl import Workbook

st.set_page_config(page_title="Project Matching App", layout="wide")

st.title("Student-Project Allocation Tool")


# ═══════════════════════════════════════════════════════
# SECTION: Template Generation
# ═══════════════════════════════════════════════════════

# --- Generate Intitial Files ---
def generate_template_excel():
    from openpyxl import Workbook
    from io import BytesIO

    wb = Workbook()

    # --- Students Sheet ---
    ws_students = wb.active
    ws_students.title = "students"
    ws_students.append(["student_id", "student_name", "choice_1", "choice_2", "choice_3"])

    # --- Projects Sheet ---
    ws_projects = wb.create_sheet("projects")
    ws_projects.append(["project_id", "project_title", "supervisor_id", "supervisor_name", "max_students"])

    # --- Supervisors Sheet ---
    ws_supervisors = wb.create_sheet("supervisors")
    ws_supervisors.append(["supervisor_id", "supervisor_name", "supervisor_email", "capacity"])

    # --- Preallocated Sheet ---
    ws_preallocated = wb.create_sheet("preallocated")
    ws_preallocated.append(["student_id", "project_id", "supervisor_id"])

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ═══════════════════════════════════════════════════════
# SECTION: Sidebar / Upload
# ═══════════════════════════════════════════════════════

# --- User Guide at top of sidebar ---
with st.sidebar.expander("📘 User Guide", expanded=False):
    st.markdown("""
    **How to Use This App**

    1. **Prepare Your Data**  
       - Use the provided template to format your input.  
       - Required sheets: `students`, `projects`, `supervisors`, and `preallocated`.

    2. **Upload Your Data**  
       - Upload a single Excel file with all sheets,  
         *or* upload the individual CSV files.

    3. **Check for Errors**  
       - Validation results will appear after upload.  
       - Fix any issues before proceeding.

    4. **Run Matching Algorithms**  
       - Algorithms: Greedy, Stable Marriage, and Linear Programming.

    5. **View and Download Results**  
       - Explore match quality, supervisor load, and satisfaction.  
       - Download results as Excel.

    Need an example? Use **“Generate Input Template”** below.
    """)

# --- Default Capacity Settings ---
st.sidebar.subheader("Default Capacity Settings")
default_supervisor_capacity = st.sidebar.number_input("Default supervisor capacity", min_value=1, value=3, step=1)
default_project_capacity = st.sidebar.number_input("Default project capacity", min_value=1, value=1, step=1)



# --- File Upload ---
st.sidebar.markdown("---")
st.sidebar.header("Upload Input Data")

# --- Template Download Button ---
st.sidebar.markdown("Need sample input files?")
if st.sidebar.button("Generate Input Template"):
    template = generate_template_excel()
    st.sidebar.download_button(
        label="Download Excel Template",
        data=template,
        file_name="input_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Option 1: Single Excel file
excel_file = st.sidebar.file_uploader("Upload Excel File with All Sheets", type="xlsx")

# Option 2: Separate CSVs
st.sidebar.markdown("Or upload individual CSV files:")
students_file = st.sidebar.file_uploader("students.csv", type="csv")
projects_file = st.sidebar.file_uploader("projects.csv", type="csv")
supervisors_file = st.sidebar.file_uploader("supervisors.csv", type="csv")
preallocated_file = st.sidebar.file_uploader("preallocated.csv", type="csv")


# ═══════════════════════════════════════════════════════
# SECTION: Data Loading
# ═══════════════════════════════════════════════════════

data_loaded = False # Flag to check if all required data has been successfully loaded

# --- Option 1: A single Excel file containing all sheets ---
if excel_file:
    try:
        # Load the Excel file and extract each sheet by name
        excel_data = pd.ExcelFile(excel_file)
        students_df = pd.read_excel(excel_data, sheet_name="students")
        projects_df = pd.read_excel(excel_data, sheet_name="projects")
        supervisors_df = pd.read_excel(excel_data, sheet_name="supervisors")
        preallocated_df = pd.read_excel(excel_data, sheet_name="preallocated")

        # Inform the user of successful upload
        st.success("Excel file loaded successfully.")
        data_loaded = True
    except Exception as e:
        # Show error message if reading fails 
        st.error(f"Failed to read Excel file: {e}")

# --- Option 2: Four separate CSV files (one per data type) ---
elif all([students_file, projects_file, supervisors_file, preallocated_file]):
    try:
        # Load each CSV into a DataFrame
        students_df = pd.read_csv(students_file)
        projects_df = pd.read_csv(projects_file)
        supervisors_df = pd.read_csv(supervisors_file)
        preallocated_df = pd.read_csv(preallocated_file)

        # Inform the user of successful upload
        st.success("CSV files loaded successfully.")
        data_loaded = True
    except Exception as e:
        # Show error message if one or more CSVs could not be read
        st.error(f"Failed to read one or more CSV files: {e}")

# --- If neither a valid Excel nor all four CSVs are provided ---
else:
    # Prompt user to upload the required input files
    st.info("⬅️ Please upload either a single Excel file or all four required CSV files to begin.")


# ═══════════════════════════════════════════════════════
# SECTION: Data Validation
# ═══════════════════════════════════════════════════════

# --- Only run logic if data loaded ---
if data_loaded:

    # Convert numeric columns (project capacity and supervisor capacity) to integers where possible
    # If invalid values are found, they are replaced with NaN to be caught in validation checks
    projects_df['max_students'] = pd.to_numeric(projects_df['max_students'], errors='coerce')
    supervisors_df['capacity'] = pd.to_numeric(supervisors_df['capacity'], errors='coerce')

    # Create lookup dictionaries for default capacities
    # If capacity is missing for a supervisor or project, use the default values defined in the sidebar
    supervisor_capacity = {
        row['supervisor_id']: int(row['capacity']) if pd.notna(row['capacity']) else default_supervisor_capacity
        for _, row in supervisors_df.iterrows()
    }

    project_capacity = {
        row['project_id']: int(row['max_students']) if pd.notna(row['max_students']) else default_project_capacity
        for _, row in projects_df.iterrows()
    }

    # --- Data Validation ---
    def validate_student_data(students_df, projects_df):
        errors = []
        required_fields = ['student_id', 'student_name', 'choice_1', 'choice_2', 'choice_3']
        missing_fields = [field for field in required_fields if field not in students_df.columns]
        if missing_fields:
            errors.append(f"Missing required columns: {missing_fields}")

        # Check for missing values
        if students_df[required_fields].isnull().any().any():
            null_ids = students_df[students_df[required_fields].isnull().any(axis=1)]['student_id'].tolist()
            errors.append(f"Missing required fields for student_ids: {null_ids}")

        # Check for duplicate IDs
        if students_df['student_id'].duplicated().any():
            dup_ids = students_df[students_df['student_id'].duplicated()]['student_id'].tolist()
            errors.append(f"Duplicate student IDs: {dup_ids}")

        # Check for invalid or duplicate project choices
        valid_project_ids = set(projects_df['project_id'])
        for _, row in students_df.iterrows():
            sid = row['student_id']
            for col in ['choice_1', 'choice_2', 'choice_3']:
                if row[col] not in valid_project_ids:
                    errors.append(f"Student {sid}: Invalid project ID '{row[col]}' in {col}")

            if len(set([row['choice_1'], row['choice_2'], row['choice_3']])) < 3:
                errors.append(f"Student {sid}: Duplicate project choices.")

        return errors

    def validate_supervisor_diversity(students_df, projects_df):
        warnings = []
        for _, row in students_df.iterrows():
            choices = [row['choice_1'], row['choice_2'], row['choice_3']]
            sup_ids = []
            for c in choices:
                pr = projects_df[projects_df['project_id'] == c]
                if not pr.empty:
                    sup_ids.append(pr['supervisor_id'].values[0])
            if len(set(sup_ids)) < 3:
                warnings.append(row['student_id'])
        return warnings
    
    def validate_project_data(projects_df):
        errors = []
        required_columns = {'project_id', 'project_title', 'supervisor_id', 'supervisor_name', 'max_students'}
        missing_columns = required_columns - set(projects_df.columns)
        if missing_columns:
            errors.append(f"Missing required columns in projects sheet: {list(missing_columns)}")
            return errors

        if projects_df['project_id'].duplicated().any():
            dup_ids = projects_df[projects_df['project_id'].duplicated()]['project_id'].tolist()
            errors.append(f"Duplicate project IDs found: {dup_ids}")

        # Allow blanks but no negative values
        invalid_rows = projects_df[
            projects_df['max_students'].notnull() & (projects_df['max_students'] < 0)
        ]
        if not invalid_rows.empty:
            invalid_ids = invalid_rows['project_id'].tolist()
            errors.append(f"Negative max_students values for project IDs: {invalid_ids}")

        return errors
    
    def validate_supervisor_data(supervisors_df):
        errors = []
        required_columns = {'supervisor_id', 'supervisor_name', 'capacity'}
        optional_columns = {'supervisor_email'}
        missing_columns = required_columns - set(supervisors_df.columns)
        if missing_columns:
            errors.append(f"Missing required columns in supervisors sheet: {list(missing_columns)}")
            return errors

        if supervisors_df['supervisor_id'].duplicated().any():
            dup_ids = supervisors_df[supervisors_df['supervisor_id'].duplicated()]['supervisor_id'].tolist()
            errors.append(f"Duplicate supervisor IDs found: {dup_ids}")

        # Allow blanks but no negative values
        invalid_rows = supervisors_df[
            supervisors_df['capacity'].notnull() & (supervisors_df['capacity'] < 0)
        ]
        if not invalid_rows.empty:
            invalid_ids = invalid_rows['supervisor_id'].tolist()
            errors.append(f"Negative capacity values for supervisor IDs: {invalid_ids}")

        return errors

    # Run all validation checks
    project_errors = validate_project_data(projects_df)
    supervisor_errors = validate_supervisor_data(supervisors_df)
    student_errors = validate_student_data(students_df, projects_df)
    diversity_warnings = validate_supervisor_diversity(students_df, projects_df)

    if student_errors or project_errors or supervisor_errors:
        st.error("Data validation failed:")
        if student_errors:
            st.subheader("Student Data Issues")
            for err in student_errors:
                st.write("-", err)
        if project_errors:
            st.subheader("Project Data Issues")
            for err in project_errors:
                st.write("-", err)
        if supervisor_errors:
            st.subheader("Supervisor Data Issues")
            for err in supervisor_errors:
                st.write("-", err)
        st.stop() # Prevents proceeding to allocation if data is invalid

    # Display warnings for non-critical issues
    if diversity_warnings:
        st.warning("Students with limited supervisor diversity in their choices:")
        st.write(diversity_warnings)

    # If no critical errors found
    st.success("Data validation passed!")


# ═══════════════════════════════════════════════════════
# SECTION: Capacity Check
# ═══════════════════════════════════════════════════════
    # Calculate total available supervisor capacity
    total_supervisor_capacity = supervisors_df['capacity'].fillna(default_supervisor_capacity).sum()

    # Calculate total available project capacity
    total_project_capacity = projects_df['max_students'].fillna(default_project_capacity).sum()

    # Count total number of students
    num_students = len(students_df)

    # Display warnings if total capacity is insufficient
    if total_supervisor_capacity < num_students:
        st.warning(f"⚠️ Total supervisor capacity ({int(total_supervisor_capacity)}) "
                f"is less than the number of students ({num_students}). "
                "Some students cannot be allocated.")
    elif total_project_capacity < num_students:
        st.warning(f"⚠️ Total project capacity ({int(total_project_capacity)}) "
                f"is less than the number of students ({num_students}). "
                "Some students cannot be allocated.")
    else:
        st.info("✅ Capacity check passed: there should be enough space for all students.")



# ═══════════════════════════════════════════════════════
# SECTION: Matching Algorithms
# ═══════════════════════════════════════════════════════

    # --- Matching Algorithms ---
    def greedy_matching(students_df, projects_df, supervisor_capacity, project_capacity, preallocated_df):
        """
        Approach:
            - Iterates through students (randomized order) and assigns each to the
            highest-ranked project on their preference list that still has available capacity.
            - Does not guarantee optimality but is fast and simple.
        """
        allocation = {}
        supervisor_load = {sup_id: 0 for sup_id in supervisor_capacity}
        project_load = {pid: 0 for pid in projects_df['project_id']}

        # Start by assigning preallocated students
        for _, row in preallocated_df.iterrows():
            sid, pid, sup = row['student_id'], row['project_id'], row['supervisor_id']
            allocation[sid] = pid
            supervisor_load[sup] += 1
            project_load[pid] += 1

        # Randomize remaining students to avoid bias in order
        students_df_shuffled = students_df.sample(frac=1).reset_index(drop=True)

        # Try to allocate each student to their top available choice
        for _, row in students_df_shuffled.iterrows():
            sid = row['student_id']
            if sid in allocation:
                continue
            for choice in ['choice_1', 'choice_2', 'choice_3']:
                pid = row[choice]
                project_row = projects_df[projects_df['project_id'] == pid]
                if project_row.empty: continue
                sup = project_row['supervisor_id'].values[0]
                max_cap = project_capacity[pid]
                # Check both supervisor and project capacity limits
                if supervisor_load[sup] < supervisor_capacity[sup] and (max_cap is None or project_load[pid] < max_cap):
                    allocation[sid] = pid
                    supervisor_load[sup] += 1
                    project_load[pid] += 1
                    break # Stop after assigning first available choice
        return allocation

    # --- Stable Marriage Algorithm ---
    def stable_marriage_matching(students_df, projects_df, supervisor_capacity, project_capacity, preallocated_df):
        """
        Approach:
            - Inspired by the Gale-Shapley "Stable Marriage" algorithm.
            - Students propose to their most preferred project that has remaining capacity.
            - If rejected (capacity full), they move on to their next choice.
            - Attempts to avoid unstable pairings but does not guarantee optimality
            if constraints are tight.
        """
        allocation = {}
        supervisor_load = {sup_id: 0 for sup_id in supervisor_capacity}
        project_load = {pid: 0 for pid in projects_df['project_id']}

        # Assign preallocated students first
        for _, row in preallocated_df.iterrows():
            sid, pid, sup = row['student_id'], row['project_id'], row['supervisor_id']
            allocation[sid] = pid
            supervisor_load[sup] += 1
            project_load[pid] += 1

        # Prepare a queue of students and their preference lists
        student_prefs = {
            row['student_id']: deque([row['choice_1'], row['choice_2'], row['choice_3']])
            for _, row in students_df.iterrows() if row['student_id'] not in allocation
        }
        free_students = deque(student_prefs.keys())

        # Iteratively assign students to available projects
        while free_students:
            sid = free_students.popleft()
            if not student_prefs[sid]:
                continue
            pid = student_prefs[sid].popleft()
            project_row = projects_df[projects_df['project_id'] == pid]
            if project_row.empty: continue
            sup = project_row['supervisor_id'].values[0]
            max_cap = project_capacity.get(pid)
            # If capacity is available, allocate; otherwise, requeue student
            if supervisor_load[sup] < supervisor_capacity[sup] and (max_cap is None or project_load[pid] < max_cap):
                allocation[sid] = pid
                supervisor_load[sup] += 1
                project_load[pid] += 1
            else:
                free_students.append(sid)
        return allocation

    # --- Linear Programming Matching Algorithm ---
    def linear_programming_matching(students_df, projects_df, supervisor_capacity, project_capacity, preallocated_df):
        """
        Approach:
            - Formulates the allocation problem as a binary integer programming model.
            - Objective: maximize total satisfaction score across all students
            (3 points for 1st choice, 2 for 2nd, 1 for 3rd).
            - Constraints:
                1) Each student assigned to at most one project
                2) Respect preallocated assignments
                3) Do not exceed project capacity
                4) Do not exceed supervisor capacity
            - Uses the PuLP solver to find the optimal allocation.
        """
        prob = LpProblem("Student_Project_Matching", LpMaximize)
        students = list(students_df['student_id'])
        projects = list(projects_df['project_id'])
        project_supervisors = projects_df.set_index('project_id')['supervisor_id'].to_dict()

        # Define satisfaction scores for each student's preferences
        student_choices = {
            row['student_id']: {
                row['choice_1']: 3,
                row['choice_2']: 2,
                row['choice_3']: 1
            } for _, row in students_df.iterrows()
        }

        # Create binary decision variables x[(student, project)]
        x = LpVariable.dicts("assign", [(s, p) for s in students for p in projects], cat=LpBinary)

        # Objective: maximize sum of satisfaction scores
        prob += lpSum(x[(s, p)] * student_choices.get(s, {}).get(p, 0) for s in students for p in projects)

        # Constraint 1: Each student assigned to at most one project
        for s in students:
            prob += lpSum(x[(s, p)] for p in projects) <= 1

        # Constraint 2: Respect preallocated students
        for _, row in preallocated_df.iterrows():
            sid, pid = row['student_id'], row['project_id']
            for p in projects:
                prob += x[(sid, p)] == int(p == pid)

        # Constraint 3: Do not exceed project capacity
        for p in projects:
            max_cap = project_capacity.get(p)
            if max_cap is not None:
                prob += lpSum(x[(s, p)] for s in students) <= max_cap

        # Constraint 4: Do not exceed supervisor capacity
        for sup, sup_cap in supervisor_capacity.items():
            sup_projects = [p for p, s in project_supervisors.items() if s == sup]
            prob += lpSum(x[(s, p)] for s in students for p in sup_projects) <= sup_cap

        # Solve optimization problem
        prob.solve(PULP_CBC_CMD(msg=False))

        # Build final allocation dictionary from solution variables
        allocation = {
            s: p for s in students for p in projects
            if x[(s, p)].varValue is not None and x[(s, p)].varValue > 0.5
        }
        return allocation

    # --- Run all matchings ---
    with st.spinner("Running matching algorithms..."):
        greedy = greedy_matching(students_df, projects_df, supervisor_capacity, project_capacity, preallocated_df)
        stable = stable_marriage_matching(students_df, projects_df, supervisor_capacity, project_capacity, preallocated_df)
        lp = linear_programming_matching(students_df, projects_df, supervisor_capacity, project_capacity, preallocated_df)

    st.success("Matching complete!")


# ═══════════════════════════════════════════════════════
# SECTION: Excel Export
# ═══════════════════════════════════════════════════════

    # Define function to create an Excel file from allocation results
    def export_excel(allocations):
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from io import BytesIO

        # Create a temporary in-memory file for the Excel output
        output = BytesIO()
        wb = Workbook()

        # Pre-compute lookup dictionaries for fast reference
        project_lookup = projects_df.set_index('project_id')[['project_title', 'supervisor_id']].to_dict('index')
        supervisor_lookup = supervisors_df.set_index('supervisor_id')['supervisor_name'].to_dict()
        supervisor_email_lookup = supervisors_df.set_index('supervisor_id')['supervisor_email'].to_dict()

        # Iterate over each allocation result (one per algorithm)
        for name, alloc in allocations.items():
            # Create a new worksheet named after the algorithm
            ws = wb.create_sheet(title=name)
            data = []

            # Build a list of student allocation details
            for _, row in students_df.iterrows():
                sid = row['student_id']
                sname = row['student_name']
                assigned_pid = alloc.get(sid, None) # Get assigned project for this student

                # Determine which preference was matched
                if assigned_pid == row['choice_1']:
                    choice_rank = "1st"
                elif assigned_pid == row['choice_2']:
                    choice_rank = "2nd"
                elif assigned_pid == row['choice_3']:
                    choice_rank = "3rd"
                elif assigned_pid is not None:
                    choice_rank = "Outside Top 3" # Assigned project not in student’s choices
                else:
                    choice_rank = "Unassigned"

                # If student was assigned, retrieve project & supervisor details
                if assigned_pid:
                    project_info = project_lookup.get(assigned_pid, {})
                    project_name = project_info.get('project_title', 'Unknown')
                    supervisor_id = project_info.get('supervisor_id', 'Unknown')
                    supervisor_name = supervisor_lookup.get(supervisor_id, 'Unknown')
                else:
                    # If unassigned, fill fields with placeholders
                    assigned_pid = 'UNASSIGNED'
                    project_name = ''
                    supervisor_name = ''
                    supervisor_id = None

                # Look up supervisor email if available
                supervisor_email = supervisor_email_lookup.get(supervisor_id, 'Unknown')

                # Append row data for this student
                data.append({
                    "student_id": sid,
                    "student_name": sname,
                    "assigned_project_id": assigned_pid,
                    "assigned_project_name": project_name,
                    "supervisor_name": supervisor_name,
                    "supervisor_email": supervisor_email,
                    "assigned_choice": choice_rank
                })

            # Convert list of dictionaries to DataFrame
            df = pd.DataFrame(data)

            # Write DataFrame to Excel worksheet row by row
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

        # Remove default empty sheet automatically created by openpyxl
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Save workbook to the in-memory BytesIO stream
        wb.save(output)
        output.seek(0) # Reset pointer to start
        return output

    # Generate downloadable Excel file containing all three algorithm results
    excel_data = export_excel({
        "Greedy": greedy,
        "Stable Marriage": stable,
        "Linear Programming": lp
    })

    # Streamlit button for user to download the file
    st.download_button("Download Excel Results", data=excel_data, file_name="matchings.xlsx")

    # Section header for the next part of the UI
    st.subheader("Match Summary")


# ═══════════════════════════════════════════════════════
# SECTION: Summary + Analysis
# ═══════════════════════════════════════════════════════

    # --- Results Overview ---
    def summarize(allocation, method):
        total = len(students_df) # Total number of students
        matched = len(allocation) # Number of students who got any project
        scores = [] # Stores satisfaction scores for each student
        choice_counts = {1: 0, 2: 0, 3: 0} # Count how many students got 1st, 2nd, 3rd choice

        # Iterate through all students and calculate their scores
        for _, row in students_df.iterrows():
            sid = row['student_id']
            pid = allocation.get(sid) # Project assigned to this student

            # Assign score based on which choice the student received
            if pid == row['choice_1']:
                scores.append(3)
                choice_counts[1] += 1
            elif pid == row['choice_2']:
                scores.append(2)
                choice_counts[2] += 1
            elif pid == row['choice_3']:
                scores.append(1)
                choice_counts[3] += 1
            elif pid is not None:
                scores.append(0) # Assigned a project outside top 3 choices
            else:
                scores.append(0) # Student remained unmatched

        return {
            "Method": method,
            "Matched": matched,
            "Unmatched": total - matched,
            "Avg Score": round(np.mean(scores), 2),
            "1st Choice": choice_counts[1],
            "2nd Choice": choice_counts[2],
            "3rd Choice": choice_counts[3],
        }

    # --- Summary Table ---
    summary_df = pd.DataFrame([
        summarize(greedy, "Greedy"),
        summarize(stable, "Stable Marriage"),
        summarize(lp, "Linear Programming")
    ])

    # Display results as a table in Streamlit
    st.dataframe(summary_df, use_container_width=True)

    # --- Match Quality Analysis ---
    def analyze_match_quality(allocation, students_df, projects_df, supervisors_df, method_name="Method"):
        st.subheader(f"Match Quality Analysis – {method_name}")

        # 1. Count how many students got 1st, 2nd, 3rd choice, or unmatched
        distribution = {'choice_1': 0, 'choice_2': 0, 'choice_3': 0, 'unmatched': 0, 'other': 0}
        for _, row in students_df.iterrows():
            sid = row['student_id']
            assigned = allocation.get(sid, None)
            if assigned is None:
                distribution['unmatched'] += 1
            elif assigned == row['choice_1']:
                distribution['choice_1'] += 1
            elif assigned == row['choice_2']:
                distribution['choice_2'] += 1
            elif assigned == row['choice_3']:
                distribution['choice_3'] += 1
            else:
                distribution['other'] += 1  # Project outside top 3

        st.markdown("**Choice Preference Distribution:**")
        choice_df = pd.DataFrame.from_dict(distribution, orient="index", columns=["Count"])
        choice_df.index.name = "Choice"
        st.dataframe(choice_df)

        # 2. Supervisor load: how many students each supervisor was assigned
        proj_to_sup = projects_df.set_index('project_id')['supervisor_id'].to_dict()
        supervisor_load = {sup: 0 for sup in supervisors_df['supervisor_id']}
        for sid, pid in allocation.items():
            sup = proj_to_sup.get(pid)
            if sup is not None:
                supervisor_load[sup] += 1

        # Replace supervisor IDs with their names for better readability
        supervisor_name_lookup = supervisors_df.set_index('supervisor_id')['supervisor_name'].to_dict()

        # Reformat dictionary to use supervisor names as keys
        supervisor_named_load = {
            supervisor_name_lookup.get(sup_id, sup_id): load
            for sup_id, load in supervisor_load.items()
        }

        # Compute basic statistics
        loads = np.array(list(supervisor_named_load.values()))
        st.markdown("**Supervisor Load Distribution:**")
        st.write(f"Min load: {loads.min()} | Max load: {loads.max()} | Mean: {loads.mean():.2f} | Std Dev: {loads.std():.2f}")

        # Create DataFrame with supervisor names as index
        sup_load_df = pd.DataFrame.from_dict(supervisor_named_load, orient='index', columns=['Students Assigned'])
        sup_load_df.index.name = "Supervisor Name"
        sup_load_df = sup_load_df.sort_values(by="Students Assigned", ascending=False)
        st.dataframe(sup_load_df)

        # 3. List students who were not assigned any of their top 3 choices
        outside = []
        for _, row in students_df.iterrows():
            sid = row['student_id']
            assigned = allocation.get(sid, None)
            if assigned is not None and assigned not in [row['choice_1'], row['choice_2'], row['choice_3']]:
                outside.append(sid)

        st.markdown("**Assigned Outside Top 3 Choices:**")
        st.write(f"{len(outside)} student(s) assigned outside their top 3.")
        if outside:
            outside_df = pd.DataFrame(outside, columns=["Student ID"])
            st.dataframe(outside_df, use_container_width=True)

        # 4. Count how many students were assigned to each project
        usage = {pid: 0 for pid in projects_df['project_id']}
        for pid in allocation.values():
            if pid in usage:
                usage[pid] += 1

        # Replace project IDs with project titles for clarity
        project_name_lookup = projects_df.set_index('project_id')['project_title'].to_dict()
        usage_named = {
            project_name_lookup.get(pid, pid): count
            for pid, count in usage.items()
        }
        st.markdown("**Project Utilization:**")
        usage_df = pd.DataFrame.from_dict(usage_named, orient='index', columns=["Assigned Count"])
        usage_df.index.name = "Project Title"
        usage_df = usage_df.sort_values(by="Assigned Count", ascending=False)
        st.dataframe(usage_df)

    # Show expandable analysis section for each algorithm
    st.subheader("Analysis")
    with st.expander("View Greedy Matching Analysis"):
        analyze_match_quality(greedy, students_df, projects_df, supervisors_df, "Greedy Matching")
    with st.expander("View Stable Marriage Analysis"):
        analyze_match_quality(stable, students_df, projects_df, supervisors_df, "Stable Marriage")
    with st.expander("View Linear Programming Analysis"):
        analyze_match_quality(lp, students_df, projects_df, supervisors_df, "Linear Programming")


# ═══════════════════════════════════════════════════════
# SECTION: Satisfaction Charts
# ═══════════════════════════════════════════════════════

    # --- Satisfaction Analysis ---
    def compute_satisfaction_scores(allocation, students_df, method_name="Method"):
        st.subheader(f"Student Satisfaction Score – {method_name}")

        # Define point system for satisfaction based on ranking
        score_weights = {'choice_1': 3, 'choice_2': 2, 'choice_3': 1}
        scores = [] # List of satisfaction scores per student
        unmatched = 0 # Count of students who did not receive any project

        # Iterate over each student and calculate their individual satisfaction score
        for _, row in students_df.iterrows():
            sid = row['student_id']
            assigned = allocation.get(sid) # Project assigned to this student

            if assigned == row['choice_1']:
                scores.append(score_weights['choice_1'])
            elif assigned == row['choice_2']:
                scores.append(score_weights['choice_2'])
            elif assigned == row['choice_3']:
                scores.append(score_weights['choice_3'])
            elif assigned is None:
                # Student was left unmatched
                unmatched += 1
                scores.append(0)
            else:
                # Student was assigned a project outside their top 3
                scores.append(0)

        # Calculate average satisfaction score across all students
        avg_score = np.mean(scores)

        # Display satisfaction metrics in the Streamlit UI
        st.markdown(f"**Average Satisfaction Score:** {avg_score:.2f}")
        st.markdown(f"**Unmatched Students:** {unmatched}")

        # Plot histogram showing distribution of satisfaction scores
        fig, ax = plt.subplots(figsize=(6, 4))
        ax.hist(scores, bins=[-0.5, 0.5, 1.5, 2.5, 3.5], edgecolor='black', align='mid', rwidth=0.8)
        ax.set_xticks([0, 1, 2, 3])
        ax.set_xlabel("Satisfaction Score (0–3)")
        ax.set_ylabel("Number of Students")
        ax.set_title(f"Satisfaction Score Distribution – {method_name}")
        ax.grid(axis='y', linestyle='--', alpha=0.7)
        st.pyplot(fig)

    st.subheader("Satisfaction")

    # Brief explanation of the scoring system
    st.markdown("**Definition:** Each student is awarded **3 points** if matched to their 1st choice, "
                "**2 points** for 2nd choice, **1 point** for 3rd choice, and **0 points** otherwise.")
    
    # Allow the user to expand/collapse each algorithm's satisfaction analysis
    with st.expander("Satisfaction – Greedy Matching"):
        compute_satisfaction_scores(greedy, students_df, "Greedy Matching")
    with st.expander("Satisfaction – Stable Marriage"):
        compute_satisfaction_scores(stable, students_df, "Stable Marriage")
    with st.expander("Satisfaction – Linear Programming"):
        compute_satisfaction_scores(lp, students_df, "Linear Programming")


# ═══════════════════════════════════════════════════════
# SECTION: Supervisor Load
# ═══════════════════════════════════════════════════════

    # --- Supervisor Load Analysis ---
    def analyze_supervisor_load(allocation, projects_df, supervisors_df, method_name="Method"):
        st.subheader(f"Supervisor Load Analysis – {method_name}")

        # Map each project to its supervisor ID for easy lookups
        proj_to_sup = projects_df.set_index('project_id')['supervisor_id'].to_dict()

        # Dictionary of supervisor capacities {supervisor_id: capacity}
        sup_capacities = supervisors_df.set_index('supervisor_id')['capacity'].to_dict()

        # Initialize a count of students assigned to each supervisor (start at zero)
        supervisor_load = {sup: 0 for sup in sup_capacities}

        # Count how many students are assigned to each supervisor based on the allocation
        for pid in allocation.values():
            sup = proj_to_sup.get(pid)
            if sup in supervisor_load:
                supervisor_load[sup] += 1

        # Prepare a list to store summarized results for display
        results = []
        for sup_id, load in supervisor_load.items():
            # Retrieve supervisor capacity (default to 0 if missing)
            cap = sup_capacities.get(sup_id, 0)

            # Retrieve supervisor name for better readability
            name = supervisors_df.loc[supervisors_df['supervisor_id'] == sup_id, 'supervisor_name'].values[0]

            # Determine status of supervisor load compared to capacity
            if load > cap:
                status = "OVERLOADED"
            elif load < cap:
                status = "UNDERUSED"
            else:
                status = "OPTIMAL"

            # Append the supervisor's load info to the results list
            results.append({
                "Supervisor Name": name,
                "Supervisor ID": sup_id,
                "Assigned": load,
                "Capacity": cap,
                "Status": status
            })

        # Convert the results list into a DataFrame for easy display
        df = pd.DataFrame(results)

        # Show the final supervisor load table in the Streamlit interface
        st.dataframe(df)

    st.subheader("Supervisor Load Analysis")
    with st.expander("Greedy Matching – Supervisor Load"):
        analyze_supervisor_load(greedy, projects_df, supervisors_df, "Greedy")
    with st.expander("Stable Marriage – Supervisor Load"):
        analyze_supervisor_load(stable, projects_df, supervisors_df, "Stable Marriage")
    with st.expander("Linear Programming – Supervisor Load"):
        analyze_supervisor_load(lp, projects_df, supervisors_df, "Linear Programming")


# ═══════════════════════════════════════════════════════
# SECTION: Project Popularity
# ═══════════════════════════════════════════════════════

    # --- Project Popularity Analysis ---
    def analyze_project_popularity_and_utilization(students_df, projects_df, allocation, method_name="Method"):
        st.subheader(f"Project Popularity & Utilization – {method_name}")

        # Initialize a dictionary for all projects with zero requests
        popularity = {pid: 0 for pid in projects_df['project_id']}
        # Loop through every student and check their 1st, 2nd, and 3rd choices
        for _, row in students_df.iterrows():
            for col in ['choice_1', 'choice_2', 'choice_3']:
                pid = row[col]
                if pid in popularity:
                    popularity[pid] += 1

        # Initialize a dictionary for all projects with zero assignments
        utilization = {pid: 0 for pid in projects_df['project_id']}
        # Loop through all final allocations and count project usage
        for pid in allocation.values():
            if pid in utilization:
                utilization[pid] += 1

         # Map project IDs to human-readable project titles
        project_id_to_title = projects_df.set_index("project_id")["project_title"].to_dict()
        data = []

        # For each project, calculate its request count, assignment count, and status
        for pid in projects_df['project_id']:
            requested = popularity.get(pid, 0)
            assigned = utilization.get(pid, 0)

            # Determine project status based on how it was allocated
            if requested == 0:
                status = "NEVER PICKED"     # No student selected this project
            elif assigned == 0:
                status = "NEVER ASSIGNED"   # Requested by students, but no one assigned
            elif assigned > requested:
                status = "OVER-ASSIGNED"    # More students assigned than requested
            elif assigned < requested:
                status = "UNDER-ASSIGNED"   # Fewer students assigned than requested
            else:
                status = "MATCHED"          # Perfect match: assigned = requested

            # Append project details to the summary list
            data.append({
                "Project Title": project_id_to_title.get(pid, "Unknown"),
                "Requested": requested,
                "Assigned": assigned,
                "Status": status
            })

        # Convert data to DataFrame for display
        df = pd.DataFrame(data)

        # Sort the table in descending order of assignments (most utilized first)
        df = df.sort_values(by="Assigned", ascending=False)

        # Use project titles as the index for cleaner display
        df.set_index("Project Title", inplace=True)

        # Display the results table in the Streamlit app
        st.dataframe(df, use_container_width=True)

    # Display results for all three algorithms
    st.subheader("Project Popularity & Utilization")
    with st.expander("Greedy Matching – Project Popularity"):
        analyze_project_popularity_and_utilization(students_df, projects_df, greedy, "Greedy")
    with st.expander("Stable Marriage – Project Popularity"):
        analyze_project_popularity_and_utilization(students_df, projects_df, stable, "Stable Marriage")
    with st.expander("Linear Programming – Project Popularity"):
        analyze_project_popularity_and_utilization(students_df, projects_df, lp, "Linear Programming")


# ═══════════════════════════════════════════════════════
# SECTION: Combined Algorithm Comparison (Charts)
# ═══════════════════════════════════════════════════════

    st.subheader("Combined Comparison of Algorithms")

    # --- Helper functions ---
    def get_choice_percentages(allocation):
        total_students = len(students_df)
        counts = {"1st": 0, "2nd": 0, "3rd": 0, "Unmatched": 0}

        # Check each student's assigned project and increment relevant choice bucket
        for _, row in students_df.iterrows():
            pid = allocation.get(row['student_id'])
            if pid == row['choice_1']:
                counts["1st"] += 1
            elif pid == row['choice_2']:
                counts["2nd"] += 1
            elif pid == row['choice_3']:
                counts["3rd"] += 1
            else:
                counts["Unmatched"] += 1
        
        # Convert raw counts into percentages for standardized comparison
        return {k: (v / total_students) * 100 for k, v in counts.items()}

    def avg_satisfaction(allocation):
        scores = []
        for _, row in students_df.iterrows():
            pid = allocation.get(row['student_id'])
            if pid == row['choice_1']:
                scores.append(3)
            elif pid == row['choice_2']:
                scores.append(2)
            elif pid == row['choice_3']:
                scores.append(1)
            else:
                scores.append(0)
        return np.mean(scores) # Calculate average score

    # --- Prepare data for charts ---
    # Create a dictionary containing choice percentages for each algorithm
    choice_data = {
        "Greedy": get_choice_percentages(greedy),
        "Stable Marriage": get_choice_percentages(stable),
        "Linear Programming": get_choice_percentages(lp)
    }
    choice_df = pd.DataFrame(choice_data).T # Convert to DataFrame (algorithms as rows)

    # Create a dictionary with average satisfaction scores per algorithm
    satisfaction_data = {
        "Greedy": avg_satisfaction(greedy),
        "Stable Marriage": avg_satisfaction(stable),
        "Linear Programming": avg_satisfaction(lp)
    }

    # --- Stacked Bar Chart: Choice Distribution (Percentages) ---
    with st.expander("Choice Distribution Comparison Across Algorithms"):
        fig1, ax1 = plt.subplots(figsize=(8, 5))
        choice_df.plot(kind='bar', stacked=True, ax=ax1)
        ax1.set_title("Choice Distribution (%) by Algorithm")
        ax1.set_xlabel("Algorithm")
        ax1.set_ylabel("Percentage of Students (%)")
        ax1.legend(title="Choice Level")
        st.pyplot(fig1)

    # --- Grouped Bar Chart: Satisfaction Scores ---
    with st.expander("Average Satisfaction Score Comparison"):
        fig2, ax2 = plt.subplots(figsize=(6, 4))
        ax2.bar(satisfaction_data.keys(), satisfaction_data.values(),
                color=['#1f77b4', '#ff7f0e', '#2ca02c'])
        ax2.set_title("Average Satisfaction Score Comparison")
        ax2.set_ylabel("Average Score (0-3)")
        ax2.set_ylim(0, 3)
        st.pyplot(fig2)




